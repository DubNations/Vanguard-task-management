"""
导出生成器 — 支持 Excel / CSV / PDF。
"""
import csv
import io
import os
from datetime import datetime
from django.core.files.base import ContentFile


def generate_export(job):
    """根据格式生成导出文件。"""
    from apps.tasks.models import Task

    # 构建查询
    qs = Task.objects.select_related('assignee', 'creator', 'reviewer').all()
    filters = job.filters or {}

    if 'status' in filters:
        qs = qs.filter(status=filters['status'])
    if 'priority' in filters:
        qs = qs.filter(priority=filters['priority'])
    if 'assignee' in filters:
        qs = qs.filter(assignee_id=filters['assignee'])

    qs = qs.order_by('-created_at')

    row_count = qs.count()

    if job.format == 'EXCEL':
        result = _export_excel(qs, job)
    elif job.format == 'CSV':
        result = _export_csv(qs, job)
    elif job.format == 'PDF':
        result = _export_pdf(qs, job)
    else:
        raise ValueError(f'不支持的导出格式: {job.format}')

    result['row_count'] = row_count
    return result


def _export_excel(qs, job):
    """导出 Excel。"""
    from apps.tasks.models import Task
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Cache count before iteration to avoid double evaluation
    row_count = qs.count()

    wb = Workbook()
    ws = wb.active
    ws.title = '任务列表'

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = ['任务编号', '标题', '状态', '优先级', '进度', '负责人', '创建人', '截止日期', '创建时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行
    status_map = dict(Task.Status.choices)
    priority_map = dict(Task.Priority.choices)

    for row, task in enumerate(qs, 2):
        data = [
            task.task_no,
            task.title,
            status_map.get(task.status, task.status),
            priority_map.get(task.priority, task.priority),
            f'{task.progress}%',
            task.assignee.username if task.assignee else '',
            task.creator.username if task.creator else '',
            task.deadline.strftime('%Y-%m-%d %H:%M') if task.deadline else '',
            task.created_at.strftime('%Y-%m-%d %H:%M'),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # 列宽
    widths = [18, 40, 10, 8, 8, 12, 12, 18, 18]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'任务导出_{timestamp}.xlsx'

    return {
        'file': ContentFile(buffer.read(), name=filename),
        'file_name': filename,
    }


def _export_csv(qs, job):
    """导出 CSV。"""
    from apps.tasks.models import Task
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    headers = ['任务编号', '标题', '状态', '优先级', '进度', '负责人', '创建人', '截止日期', '创建时间']
    writer.writerow(headers)

    status_map = dict(Task.Status.choices)
    priority_map = dict(Task.Priority.choices)

    for task in qs:
        writer.writerow([
            task.task_no,
            task.title,
            status_map.get(task.status, task.status),
            priority_map.get(task.priority, task.priority),
            f'{task.progress}%',
            task.assignee.username if task.assignee else '',
            task.creator.username if task.creator else '',
            task.deadline.strftime('%Y-%m-%d %H:%M') if task.deadline else '',
            task.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'任务导出_{timestamp}.csv'
    content = buffer.getvalue().encode('utf-8-sig')

    return {
        'file': ContentFile(content, name=filename),
        'file_name': filename,
    }


def _export_pdf(qs, job):
    """导出 PDF（使用 WeasyPrint，本地开发环境不可用时优雅降级）。"""
    from apps.tasks.models import Task

    try:
        from weasyprint import HTML
    except ImportError:
        raise ValueError(
            'PDF 导出在本地开发环境不可用（缺少 WeasyPrint），请使用 Excel 或 CSV 格式导出。'
        )

    from django.template.loader import render_to_string

    context = {
        'tasks': qs,
        'generated_at': datetime.now(),
        'status_map': dict(Task.Status.choices),
        'priority_map': dict(Task.Priority.choices),
    }

    html_string = render_to_string('exports/task_report.html', context)
    pdf_bytes = HTML(string=html_string).write_pdf()

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'任务报表_{timestamp}.pdf'

    return {
        'file': ContentFile(pdf_bytes, name=filename),
        'file_name': filename,
    }
