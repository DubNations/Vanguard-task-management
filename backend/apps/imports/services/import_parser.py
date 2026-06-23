"""
导入文件解析器 — 支持 xlsx/csv/wps（含旧版 WPS）。
"""
import csv
import io
import os
import subprocess
import zipfile
import shutil
import tempfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook


# 标准列名 → 字段映射
DEFAULT_COLUMN_MAPPING = {
    # 旧模板
    '标题': 'title',
    '任务标题': 'title',
    'task': 'title',
    'title': 'title',
    '负责人': 'assignee_name',
    '执行人': 'assignee_name',
    'assignee': 'assignee_name',
    '截止日期': 'deadline',
    '到期日': 'deadline',
    'deadline': 'deadline',
    '状态': 'status',
    'status': 'status',
    '优先级': 'priority',
    'priority': 'priority',
    '进度': 'progress',
    'progress': 'progress',
    '备注': 'description',
    '描述': 'description',
    'description': 'description',
    # 新模板（特战队任务单）
    '任务名称': 'title',
    '任务来源': 'task_source',
    '事项名称': 'title',
    '事项内容': 'content',
    '牵头人': 'lead_name',
    '参与人': 'participant_names',
    '开始时间': 'start_date',
    '结束时间': 'end_date',
    '时间周期': 'time_range',
    '积分赋分': 'reward_info',
    '任务赋分': 'reward_info',
    '核心输出成果': 'output',
    '完成标准': 'completion_criteria',
    '附件': 'attachments',
    '派发人': 'dispatcher_name',
}

# 单任务表单的 key→field 映射
FORM_KEY_MAPPING = {
    '任务名称': 'title',
    '事项名称': 'title',
    '任务来源': 'task_source',
    '事项内容': 'content',
    '牵头人': 'lead_name',
    '参与人': 'participant_names',
    '开始时间': 'start_date',
    '结束时间': 'end_date',
    '时间周期': 'time_range',
    '积分赋分': 'reward_info',
    '任务赋分': 'reward_info',
    '核心输出成果': 'output',
    '完成标准': 'completion_criteria',
    '附件': 'attachments',
    '派发人': 'dispatcher_name',
}


def parse_import_file(session):
    """解析导入文件，返回预览数据。"""
    file_path = session.file.path
    ext = session.file_format

    if ext in ('xlsx', 'xls', 'wps'):
        return _parse_xlsx(file_path, ext)
    elif ext == 'csv':
        return _parse_csv(file_path)
    else:
        raise ValueError(f'不支持的文件格式: {ext}')


def _parse_xlsx(file_path, ext='xlsx'):
    """解析 Excel / WPS 文件。WPS 先尝试直接读取（新版兼容），失败则走 libreoffice 转换。"""
    # 先尝试直接用 openpyxl 打开（xlsx 和新版 wps 兼容）
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        if ext == 'wps':
            # 旧版 .wps 格式，需要 libreoffice 转换
            return _parse_wps_via_libreoffice(file_path)
        raise

    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return _empty_result()

    # 检测是否为单任务表单格式（key-value 布局）
    if _is_form_layout(rows):
        result = _parse_form_layout(rows)
        wb.close()
        return result

    # 标准表格格式：第一行是表头
    headers = [str(h).strip() if h else '' for h in rows[0]]
    mapping = _build_mapping(headers)

    data_rows = rows[1:]
    preview = []
    errors = []
    valid = 0

    for idx, row in enumerate(data_rows, start=2):
        record = {}
        for col_idx, header in enumerate(headers):
            if col_idx < len(row) and header in mapping:
                record[mapping[header]] = str(row[col_idx]) if row[col_idx] is not None else ''

        # 验证
        row_errors = _validate_row(idx, record)
        if row_errors:
            errors.extend(row_errors)
        else:
            valid += 1

        preview.append({
            'row': idx,
            'data': record,
            'valid': len(row_errors) == 0,
        })

    wb.close()

    return {
        'total_rows': len(data_rows),
        'valid_rows': valid,
        'error_rows': len(data_rows) - valid,
        'preview_data': preview,
        'column_mapping': mapping,
        'errors': errors,
    }


def _parse_csv(file_path):
    """解析 CSV 文件。"""
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return _empty_result()

    headers = [h.strip() for h in rows[0]]
    mapping = _build_mapping(headers)

    data_rows = rows[1:]
    preview = []
    errors = []
    valid = 0

    for idx, row in enumerate(data_rows, start=2):
        record = {}
        for col_idx, header in enumerate(headers):
            if col_idx < len(row) and header in mapping:
                record[mapping[header]] = row[col_idx].strip()

        row_errors = _validate_row(idx, record)
        if row_errors:
            errors.extend(row_errors)
        else:
            valid += 1

        preview.append({
            'row': idx,
            'data': record,
            'valid': len(row_errors) == 0,
        })

    return {
        'total_rows': len(data_rows),
        'valid_rows': valid,
        'error_rows': len(data_rows) - valid,
        'preview_data': preview,
        'column_mapping': mapping,
        'errors': errors,
    }


def _parse_wps_via_libreoffice(file_path):
    """WPS 文件 — 多层策略解析。"""
    # 策略 1: ZIP 检测 + 直接 XML 解析（不依赖 openpyxl/libreoffice）
    if _is_zip_file(file_path):
        result = _parse_wps_via_xml(file_path)
        if result is not None:
            return result

        # 策略 2: ZIP 重打包为 xlsx 再用 openpyxl 解析
        result = _try_repack_wps_to_xlsx(file_path)
        if result is not None:
            return result

    # 策略 3: OLE2 格式（WPS 文字文档）→ WPS CLI 转换链
    if _is_ole2_file(file_path):
        result = _convert_wps_via_cli(file_path)
        if result is not None:
            return result

    # 策略 4: libreoffice 转换（需要安装）
    return _convert_wps_via_libreoffice(file_path)


def _is_zip_file(file_path):
    """检测文件是否为 ZIP 格式。"""
    try:
        with zipfile.ZipFile(file_path, 'r') as zf:
            zf.namelist()
        return True
    except (zipfile.BadZipFile, Exception):
        return False


def _is_ole2_file(file_path):
    """检测文件是否为 OLE2 复合文档格式（.doc/.wps 旧版）。"""
    try:
        with open(file_path, 'rb') as f:
            header = f.read(8)
        # OLE2 magic: D0 CF 11 E0 A1 B1 1A E1
        return header == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
    except Exception:
        return False


def _find_wpscli():
    """查找 wpscli / kwpsconvert.exe 路径。"""
    import glob as glob_mod
    # 常见安装路径
    candidates = [
        r'C:\Program Files (x86)\WPS\WPS Office\*\office6\kwpsconvert.exe',
        r'C:\Program Files\WPS\WPS Office\*\office6\kwpsconvert.exe',
        os.path.expanduser(r'~\AppData\Local\Kingsoft\WPS Office\*\office6\kwpsconvert.exe'),
    ]
    for pattern in candidates:
        matches = glob_mod.glob(pattern)
        if matches:
            return matches[-1]  # 取最新版本
    # 尝试 PATH
    import shutil
    found = shutil.which('kwpsconvert') or shutil.which('wpscli')
    return found


def _convert_wps_via_cli(file_path):
    """WPS 文字文档（OLE2）→ WPS CLI 转换链：.wps → .doc → .pdf → .xlsx。

    需要本机安装 WPS Office 且用户已登录 VIP 账号。
    """
    wpscli = _find_wpscli()
    if not wpscli:
        return None

    tmp_dir = tempfile.mkdtemp()
    try:
        # Step 1: .wps → .doc（重命名，WPS Writer 二进制格式兼容 .doc）
        doc_path = os.path.join(tmp_dir, 'input.doc')
        shutil.copy2(file_path, doc_path)

        # Step 2: .doc → .pdf
        pdf_path = os.path.join(tmp_dir, 'output.pdf')
        r1 = subprocess.run(
            [wpscli, 'word2pdf', doc_path, '-o', pdf_path],
            capture_output=True, timeout=120,
        )
        if r1.returncode != 0 or not os.path.exists(pdf_path):
            return None

        # Step 3: .pdf → .xlsx
        xlsx_path = os.path.join(tmp_dir, 'output.xlsx')
        r2 = subprocess.run(
            [wpscli, 'pdf2excel', pdf_path, '-o', xlsx_path],
            capture_output=True, timeout=120,
        )
        if r2.returncode != 0 or not os.path.exists(xlsx_path):
            return None

        # Step 4: 解析 xlsx
        return _parse_xlsx(xlsx_path)

    except (subprocess.TimeoutExpired, Exception):
        return None
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _parse_wps_via_xml(file_path):
    """直接解析 ZIP 内的 XML（不依赖 openpyxl）。

    支持 xlsx 和 WPS 2019+ 的 ZIP-based 格式。
    从 xl/worksheets/sheet1.xml 或类似路径提取行列数据。
    """
    try:
        with zipfile.ZipFile(file_path, 'r') as zf:
            names = zf.namelist()

            # 定位 sharedStrings.xml（单元格文本值）
            ss_path = None
            for n in names:
                if n.endswith('sharedStrings.xml') or n.endswith('sharedStrings.xml/'):
                    ss_path = n
                    break

            # 定位第一个 worksheet
            ws_path = None
            for pattern in ['xl/worksheets/sheet1.xml', 'xl/worksheets/sheet.xml',
                            'Worksheets/sheet1.xml', 'word/tables/table1.xml']:
                if pattern in names:
                    ws_path = pattern
                    break
            if ws_path is None:
                for n in names:
                    if 'sheet1.xml' in n.lower() or ('sheet' in n.lower() and n.endswith('.xml')):
                        ws_path = n
                        break
            if ws_path is None:
                return None

            # 解析 sharedStrings
            shared_strings = []
            if ss_path:
                ss_xml = zf.read(ss_path)
                ss_root = ET.fromstring(ss_xml)
                ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for si in ss_root.findall('.//ns:si', ns):
                    texts = si.findall('.//ns:t', ns)
                    shared_strings.append(''.join(t.text or '' for t in texts))

            # 解析 worksheet
            ws_xml = zf.read(ws_path)
            ws_root = ET.fromstring(ws_xml)
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            rows_data = []
            for row_el in ws_root.findall('.//ns:row', ns):
                row_cells = {}
                for c in row_el.findall('ns:c', ns):
                    ref = c.get('r', '')  # e.g. A1, B2
                    cell_type = c.get('t', '')

                    value = ''
                    if cell_type == 's':
                        # shared string reference
                        v_el = c.find('ns:v', ns)
                        if v_el is not None and v_el.text:
                            idx = int(v_el.text)
                            value = shared_strings[idx] if idx < len(shared_strings) else ''
                    elif cell_type == 'inlineStr':
                        # inline string: <is><t>text</t></is>
                        is_el = c.find('ns:is', ns)
                        if is_el is not None:
                            texts = is_el.findall('.//ns:t', ns)
                            value = ''.join(t.text or '' for t in texts)
                    else:
                        # number / date / general
                        v_el = c.find('ns:v', ns)
                        value = v_el.text if v_el is not None else ''

                    # column letter from ref
                    col_letter = ''.join(ch for ch in ref if ch.isalpha())
                    row_cells[col_letter] = value

                rows_data.append(row_cells)

            if not rows_data:
                return None

            # 确定列范围
            all_cols = set()
            for r in rows_data:
                all_cols.update(r.keys())
            if not all_cols:
                return None
            sorted_cols = sorted(all_cols, key=lambda c: (len(c), c))

            # 转换为二维数组
            rows_list = []
            for row_cells in rows_data:
                row_arr = [row_cells.get(col, '') for col in sorted_cols]
                rows_list.append(row_arr)

            if not rows_list:
                return None

            # 第一行当表头
            headers = [str(h).strip() if h else '' for h in rows_list[0]]
            mapping = _build_mapping(headers)
            data_rows = rows_list[1:]

            preview = []
            errors = []
            valid = 0
            for idx, row in enumerate(data_rows, start=2):
                record = {}
                for col_idx, header in enumerate(headers):
                    if col_idx < len(row) and header in mapping:
                        record[mapping[header]] = str(row[col_idx]) if row[col_idx] else ''
                row_errors = _validate_row(idx, record)
                if row_errors:
                    errors.extend(row_errors)
                else:
                    valid += 1
                preview.append({
                    'row': idx,
                    'data': record,
                    'valid': len(row_errors) == 0,
                })

            return {
                'total_rows': len(data_rows),
                'valid_rows': valid,
                'error_rows': len(data_rows) - valid,
                'preview_data': preview,
                'column_mapping': mapping,
                'errors': errors,
            }

    except Exception:
        return None


def _try_repack_wps_to_xlsx(file_path):
    """尝试将 ZIP-based WPS 重打包为 xlsx 并解析。

    新版 WPS (.wps) 内部结构类似 xlsx，但目录名可能不同。
    策略：直接尝试用 openpyxl 打开（某些 WPS 版本直接兼容），
    如果不行则解压后重打包。
    """
    tmp_dir = tempfile.mkdtemp()
    try:
        # 直接尝试 openpyxl（部分新版 WPS 文件可直接读取）
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            wb.close()
            # openpyxl 能打开，回退到正常流程
            return None
        except Exception:
            pass

        # 解压并尝试重打包
        extract_dir = os.path.join(tmp_dir, 'extracted')
        os.makedirs(extract_dir)
        with zipfile.ZipFile(file_path, 'r') as zf:
            zf.extractall(extract_dir)

        # 检查是否有 [Content_Types].xlsx（xlsx 的标志文件）
        content_types = os.path.join(extract_dir, '[Content_Types].xml')
        if not os.path.exists(content_types):
            return None  # 不是 xlsx 兼容格式

        # 重打包为 xlsx
        repacked_path = os.path.join(tmp_dir, 'repacked.xlsx')
        with zipfile.ZipFile(repacked_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for root, dirs, files in os.walk(extract_dir):
                for file in files:
                    file_path_inner = os.path.join(root, file)
                    arcname = os.path.relpath(file_path_inner, extract_dir)
                    zout.write(file_path_inner, arcname)

        # 尝试解析重打包后的文件
        try:
            wb = load_workbook(repacked_path, read_only=True, data_only=True)
            wb.close()
            return _parse_xlsx(repacked_path)
        except Exception:
            return None

    except Exception:
        return None
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _convert_wps_via_libreoffice(file_path):
    """通过 libreoffice 转换 WPS → xlsx。"""
    tmp_dir = tempfile.mkdtemp()
    try:
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'xlsx', '--outdir', tmp_dir, file_path],
            capture_output=True, timeout=60,
        )
        if result.returncode != 0:
            raise ValueError(
                'WPS 文件无法解析。请在 WPS Office 中将文件另存为 .xlsx 格式后重新上传。'
            )
        converted = os.path.join(tmp_dir, os.path.splitext(os.path.basename(file_path))[0] + '.xlsx')
        if not os.path.exists(converted):
            raise ValueError(
                'WPS 文件转换失败。请在 WPS Office 中将文件另存为 .xlsx 格式后重新上传。'
            )
        return _parse_xlsx(converted)
    except FileNotFoundError:
        raise ValueError(
            'WPS 文件无法解析，且服务器未安装 LibreOffice。'
            '请在 WPS Office 中将文件另存为 .xlsx 格式后重新上传。'
        )


def _is_form_layout(rows):
    """检测是否为单任务表单格式（key-value 布局）。

    表单特征：第一行是标题，后续行第一列是字段名（在 FORM_KEY_MAPPING 中）。
    """
    if len(rows) < 3:
        return False

    # 统计第一列中匹配 FORM_KEY_MAPPING 的行数
    match_count = 0
    total_data_rows = 0
    for row in rows[1:]:  # 跳过标题行
        if row and row[0]:
            key = str(row[0]).strip()
            if key in FORM_KEY_MAPPING:
                match_count += 1
            total_data_rows += 1

    # 如果超过一半的行第一列是已知字段名，则认为是表单格式
    return total_data_rows > 0 and match_count / total_data_rows > 0.5


def _parse_form_layout(rows):
    """解析单任务表单格式（key-value 布局）。

    转换为单行记录，返回标准导入预览格式。
    支持两种列布局：
    - 2列：key | value
    - 4列：key | value | key | value（并排布局）
    """
    record = {}

    for row in rows[1:]:  # 跳过标题行
        if not row:
            continue

        # 处理 2 列布局: key | value
        if row[0]:
            key = str(row[0]).strip()
            val = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if key in FORM_KEY_MAPPING:
                record[FORM_KEY_MAPPING[key]] = val

        # 处理 4 列布局: key | value | key | value
        if len(row) >= 4 and row[2]:
            key2 = str(row[2]).strip()
            val2 = str(row[3]).strip() if row[3] else ''
            if key2 in FORM_KEY_MAPPING:
                record[FORM_KEY_MAPPING[key2]] = val2

    if not record.get('title'):
        return _empty_result_with_error('未能从表单中识别到任务名称')

    # 合并 content → description（如果 description 为空）
    if record.get('content') and not record.get('description'):
        record['description'] = record.pop('content')

    # 解析开始/结束时间
    # start_date 保留为 started_at，end_date 转为 deadline
    start_str = record.pop('start_date', '')
    end_str = record.pop('end_date', '')
    if start_str:
        record['started_at'] = start_str  # 保留开始时间
    if end_str:
        record['deadline'] = end_str
    elif start_str:
        record['deadline'] = start_str

    # 提取积分数字
    reward_raw = record.pop('reward_info', '')
    if reward_raw:
        import re
        nums = re.findall(r'(\d+)', reward_raw)
        if nums:
            record['reward_points'] = nums[0]

    # 保留附件到 custom_fields（附件是非标准字段，其他模板字段由 executor 写入正式字段）
    custom = {}
    if 'attachments' in record:
        custom['attachments'] = record.pop('attachments')
    if custom:
        record['custom_fields'] = custom

    # 参与人保留原始文本
    if 'participant_names' in record:
        record['participant_names_text'] = record.pop('participant_names')

    row_errors = _validate_row(2, record)

    return {
        'total_rows': 1,
        'valid_rows': 0 if row_errors else 1,
        'error_rows': 1 if row_errors else 0,
        'preview_data': [{
            'row': 2,
            'data': record,
            'valid': len(row_errors) == 0,
        }],
        'column_mapping': {k: k for k in record.keys()},
        'errors': row_errors,
    }


def _empty_result_with_error(message):
    return {
        'total_rows': 0,
        'valid_rows': 0,
        'error_rows': 0,
        'preview_data': [],
        'column_mapping': {},
        'errors': [{'row': 0, 'message': message}],
    }


def _build_mapping(headers):
    """自动匹配列名到字段。"""
    mapping = {}
    for header in headers:
        normalized = header.lower().strip()
        if normalized in DEFAULT_COLUMN_MAPPING:
            mapping[header] = DEFAULT_COLUMN_MAPPING[normalized]
        elif header in DEFAULT_COLUMN_MAPPING:
            mapping[header] = DEFAULT_COLUMN_MAPPING[header]
    return mapping


def _validate_row(row_num, record):
    """验证单行数据。"""
    errors = []
    title = record.get('title', '').strip()
    if not title:
        errors.append({'row': row_num, 'field': 'title', 'message': '标题不能为空'})

    priority = record.get('priority', '')
    if priority and priority not in ('LOW', 'MEDIUM', 'HIGH', 'URGENT', '低', '中', '高', '紧急'):
        errors.append({'row': row_num, 'field': 'priority', 'message': f'无效的优先级: {priority}'})

    return errors


def _empty_result():
    return {
        'total_rows': 0,
        'valid_rows': 0,
        'error_rows': 0,
        'preview_data': [],
        'column_mapping': {},
        'errors': [{'row': 0, 'message': '文件为空'}],
    }
