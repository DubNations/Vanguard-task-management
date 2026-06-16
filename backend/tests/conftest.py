"""
pytest 配置和 fixtures — 全栈模拟测试增强版。
"""
import csv
import io
import os
import tempfile

import pytest
from django.test import override_settings


# ---------------------------------------------------------------------------
# 全局清理限流缓存 — 防止测试间共享缓存导致意外 429
# ---------------------------------------------------------------------------
@pytest.fixture(autouse=True)
def _clear_throttle_cache():
    """每个测试前后清理限流缓存。"""
    from django.core.cache import cache
    cache.clear()
    yield
    cache.clear()


# ---------------------------------------------------------------------------
# 团队
# ---------------------------------------------------------------------------
@pytest.fixture
def team(db):
    from apps.accounts.models import Team
    return Team.objects.create(name='尖兵一组', description='测试团队')


# ---------------------------------------------------------------------------
# 用户
# ---------------------------------------------------------------------------
@pytest.fixture
def admin_user(db):
    from apps.accounts.models import User
    return User.objects.create_superuser(
        email='test@test.com',
        password='testpass123',
        username='testadmin',
    )


@pytest.fixture
def leader_user(db, team):
    from apps.accounts.models import User
    return User.objects.create_user(
        email='leader@test.com',
        password='testpass123',
        username='testleader',
        role='LEADER',
        team=team,
    )


@pytest.fixture
def regular_user(db, team):
    from apps.accounts.models import User
    return User.objects.create_user(
        email='user@test.com',
        password='testpass123',
        username='testuser',
        role='MEMBER',
        team=team,
    )


@pytest.fixture
def member_b_user(db, team):
    from apps.accounts.models import User
    return User.objects.create_user(
        email='memberb@test.com',
        password='testpass123',
        username='memberb',
        role='MEMBER',
        team=team,
    )


# ---------------------------------------------------------------------------
# 已认证客户端 — 每个 fixture 创建独立 APIClient 避免共享状态
# ---------------------------------------------------------------------------
@pytest.fixture
def api_client():
    from rest_framework.test import APIClient
    return APIClient()


def _make_auth_client(user):
    """创建独立的已认证客户端。"""
    from rest_framework.test import APIClient
    from rest_framework_simplejwt.tokens import RefreshToken
    client = APIClient()
    refresh = RefreshToken.for_user(user)
    client.credentials(HTTP_AUTHORIZATION=f'Bearer {refresh.access_token}')
    return client


@pytest.fixture
def auth_client(admin_user):
    return _make_auth_client(admin_user)


@pytest.fixture
def leader_client(leader_user):
    return _make_auth_client(leader_user)


@pytest.fixture
def member_client(regular_user):
    return _make_auth_client(regular_user)


@pytest.fixture
def member_b_client(member_b_user):
    return _make_auth_client(member_b_user)


# ---------------------------------------------------------------------------
# 标准任务
# ---------------------------------------------------------------------------
@pytest.fixture
def sample_task(admin_user):
    from apps.tasks.services.task_service import TaskService
    return TaskService.create_task({'title': '标准测试任务'}, admin_user)


# ---------------------------------------------------------------------------
# 文件上传用临时 MEDIA_ROOT
# ---------------------------------------------------------------------------
@pytest.fixture
def media_root(tmp_path, settings):
    settings.MEDIA_ROOT = str(tmp_path)
    return tmp_path


# ---------------------------------------------------------------------------
# 导入用测试文件
# ---------------------------------------------------------------------------
@pytest.fixture
def sample_xlsx(tmp_path):
    """生成有效的 xlsx 测试文件。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['标题', '负责人', '优先级', '截止日期', '描述'])
    ws.append(['测试任务A', 'testuser', '高', '2026-12-31', '描述A'])
    ws.append(['测试任务B', 'memberb', '中', '2026-12-31', '描述B'])
    ws.append(['测试任务C', '', '低', '', '描述C'])

    path = tmp_path / 'import_test.xlsx'
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def sample_xlsx_with_errors(tmp_path):
    """生成含错误行的 xlsx 文件。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['标题', '优先级'])
    ws.append(['有效任务', '高'])
    ws.append(['', '中'])          # 空标题 -> 错误行
    ws.append(['另一有效任务', 'INVALID_PRIORITY'])  # 无效优先级 -> 错误行

    path = tmp_path / 'import_errors.xlsx'
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def sample_xlsx_all_invalid(tmp_path):
    """生成全部无效行的 xlsx 文件。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['标题', '优先级'])
    ws.append(['', '高'])
    ws.append(['', '中'])

    path = tmp_path / 'import_all_invalid.xlsx'
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def sample_csv_file(tmp_path):
    """生成有效的 csv 测试文件。"""
    path = tmp_path / 'import_test.csv'
    with open(str(path), 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['标题', '负责人', '优先级', '描述'])
        writer.writerow(['CSV任务A', 'testuser', '高', '描述A'])
        writer.writerow(['CSV任务B', '', '中', '描述B'])
    return path


@pytest.fixture
def empty_file(tmp_path):
    """空文件。"""
    path = tmp_path / 'empty.xlsx'
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def txt_file(tmp_path):
    """不支持的 .txt 文件。"""
    path = tmp_path / 'test.txt'
    path.write_text('hello world', encoding='utf-8')
    return path


# ---------------------------------------------------------------------------
# 辅助函数
# ---------------------------------------------------------------------------
def get_jwt_token(user):
    """获取用户的 access token 字符串。"""
    from rest_framework_simplejwt.tokens import RefreshToken
    refresh = RefreshToken.for_user(user)
    return str(refresh.access_token)
