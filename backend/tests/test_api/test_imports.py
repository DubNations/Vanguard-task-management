"""导入模块测试 — 两阶段导入(上传预览+确认执行)。"""
import pytest
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile

from apps.imports.models import ImportSession
from apps.tasks.models import Task, TaskHistory


@pytest.mark.django_db
class TestImportUpload:
    """上传预览测试。"""

    def test_upload_valid_xlsx(self, auth_client, sample_xlsx):
        """上传有效 xlsx → status=PREVIEW, total_rows, valid_rows, column_mapping。"""
        with open(sample_xlsx, 'rb') as f:
            response = auth_client.post(
                reverse('import-upload'),
                {'file': f},
                format='multipart'
            )
        assert response.status_code == 200
        data = response.json()
        assert data['status'] == 'PREVIEW'
        assert data['total_rows'] == 3
        assert data['valid_rows'] >= 2
        assert 'column_mapping' in data

    def test_upload_valid_csv(self, auth_client, sample_csv_file):
        """上传有效 csv → status=PREVIEW。"""
        with open(sample_csv_file, 'rb') as f:
            response = auth_client.post(
                reverse('import-upload'),
                {'file': f},
                format='multipart'
            )
        assert response.status_code == 200
        assert response.json()['status'] == 'PREVIEW'

    def test_upload_unsupported_txt(self, auth_client, txt_file):
        """上传 .txt → 400。"""
        with open(txt_file, 'rb') as f:
            response = auth_client.post(
                reverse('import-upload'),
                {'file': f},
                format='multipart'
            )
        assert response.status_code == 400
        assert '仅支持' in response.json()['error']

    def test_upload_empty_xlsx(self, auth_client, empty_file):
        """上传空 xlsx → 返回错误。"""
        with open(empty_file, 'rb') as f:
            response = auth_client.post(
                reverse('import-upload'),
                {'file': f},
                format='multipart'
            )
        assert response.status_code == 200
        data = response.json()
        assert data['total_rows'] == 0 or len(data['errors']) > 0

    def test_upload_without_file(self, auth_client):
        """不传文件 → 400。"""
        response = auth_client.post(reverse('import-upload'), {}, format='multipart')
        assert response.status_code == 400
        assert '请选择文件' in response.json()['error']

    def test_xlsx_with_invalid_rows(self, auth_client, sample_xlsx_with_errors):
        """含错误行 → error_rows > 0。"""
        with open(sample_xlsx_with_errors, 'rb') as f:
            response = auth_client.post(
                reverse('import-upload'),
                {'file': f},
                format='multipart'
            )
        assert response.status_code == 200
        data = response.json()
        assert data['error_rows'] > 0
        assert len(data['errors']) > 0
        assert 'row' in data['errors'][0]

    def test_chinese_column_mapping(self, auth_client, sample_xlsx):
        """中文列名映射。"""
        with open(sample_xlsx, 'rb') as f:
            response = auth_client.post(
                reverse('import-upload'),
                {'file': f},
                format='multipart'
            )
        data = response.json()
        mapping = data['column_mapping']
        assert '标题' in mapping or any('title' in v for v in mapping.values())

    def test_english_column_mapping(self, auth_client, tmp_path):
        """英文列名映射。"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['title', 'priority', 'assignee'])
        ws.append(['Task1', 'HIGH', 'user1'])
        path = tmp_path / 'eng.xlsx'
        wb.save(str(path))
        wb.close()

        with open(path, 'rb') as f:
            response = auth_client.post(
                reverse('import-upload'),
                {'file': f},
                format='multipart'
            )
        data = response.json()
        assert 'title' in data['column_mapping'] or any('title' in v for v in data['column_mapping'].values())


@pytest.mark.django_db
class TestImportConfirm:
    """确认导入测试。"""

    def test_confirm_preview(self, auth_client, admin_user, sample_xlsx):
        """确认 PREVIEW 状态 → 200, 任务创建。"""
        with open(sample_xlsx, 'rb') as f:
            resp = auth_client.post(reverse('import-upload'), {'file': f}, format='multipart')
        session_id = resp.json()['session_id']

        response = auth_client.post(reverse('import-confirm', kwargs={'pk': session_id}))
        assert response.status_code == 200
        data = response.json()
        assert data['imported_count'] >= 2
        assert Task.objects.filter(source__startswith='IMPORT:').exists()

    def test_confirm_completed(self, auth_client, admin_user, sample_xlsx):
        """确认已完成的会话 → 400。"""
        with open(sample_xlsx, 'rb') as f:
            resp = auth_client.post(reverse('import-upload'), {'file': f}, format='multipart')
        session_id = resp.json()['session_id']

        auth_client.post(reverse('import-confirm', kwargs={'pk': session_id}))
        response = auth_client.post(reverse('import-confirm', kwargs={'pk': session_id}))
        assert response.status_code == 400

    def test_confirm_nonexistent(self, auth_client):
        """确认不存在的会话 → 404。"""
        import uuid
        response = auth_client.post(reverse('import-confirm', kwargs={'pk': uuid.uuid4()}))
        assert response.status_code == 404

    def test_imported_task_history(self, auth_client, admin_user, sample_xlsx):
        """导入的任务有 IMPORTED 历史记录。"""
        with open(sample_xlsx, 'rb') as f:
            resp = auth_client.post(reverse('import-upload'), {'file': f}, format='multipart')
        session_id = resp.json()['session_id']

        auth_client.post(reverse('import-confirm', kwargs={'pk': session_id}))
        assert TaskHistory.objects.filter(action='IMPORTED').exists()

    def test_import_nonexistent_assignee(self, auth_client, admin_user, tmp_path):
        """导入不存在的负责人 → assignee=None。"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['标题', '负责人'])
        ws.append(['测试任务', '不存在的用户'])
        path = tmp_path / 'no_assignee.xlsx'
        wb.save(str(path))
        wb.close()

        with open(path, 'rb') as f:
            resp = auth_client.post(reverse('import-upload'), {'file': f}, format='multipart')
        session_id = resp.json()['session_id']

        response = auth_client.post(reverse('import-confirm', kwargs={'pk': session_id}))
        assert response.status_code == 200
        task = Task.objects.filter(title='测试任务').first()
        assert task is not None
        assert task.assignee is None

    def test_priority_mapping(self, auth_client, admin_user, tmp_path):
        """优先级中文映射。"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['标题', '优先级'])
        ws.append(['高优先级任务', '高'])
        path = tmp_path / 'priority.xlsx'
        wb.save(str(path))
        wb.close()

        with open(path, 'rb') as f:
            resp = auth_client.post(reverse('import-upload'), {'file': f}, format='multipart')
        session_id = resp.json()['session_id']

        auth_client.post(reverse('import-confirm', kwargs={'pk': session_id}))
        task = Task.objects.filter(title='高优先级任务').first()
        assert task.priority == 'HIGH'

    def test_status_mapping(self, auth_client, admin_user, tmp_path):
        """状态中文映射。"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['标题', '状态'])
        ws.append(['进行中任务', '进行中'])
        path = tmp_path / 'status.xlsx'
        wb.save(str(path))
        wb.close()

        with open(path, 'rb') as f:
            resp = auth_client.post(reverse('import-upload'), {'file': f}, format='multipart')
        session_id = resp.json()['session_id']

        auth_client.post(reverse('import-confirm', kwargs={'pk': session_id}))
        task = Task.objects.filter(title='进行中任务').first()
        assert task.status == 'IN_PROGRESS'


@pytest.mark.django_db
class TestImportHistory:
    """导入历史测试。"""

    def test_import_history(self, auth_client, admin_user, sample_xlsx):
        """导入历史列表。"""
        with open(sample_xlsx, 'rb') as f:
            auth_client.post(reverse('import-upload'), {'file': f}, format='multipart')

        response = auth_client.get(reverse('import-history'))
        assert response.status_code == 200
        data = response.json()
        assert len(data) >= 1
        assert 'created_at' in data[0]

    def test_session_detail(self, auth_client, admin_user, sample_xlsx):
        """会话详情。"""
        with open(sample_xlsx, 'rb') as f:
            resp = auth_client.post(reverse('import-upload'), {'file': f}, format='multipart')
        session_id = resp.json()['session_id']

        response = auth_client.get(reverse('import-detail', kwargs={'pk': session_id}))
        assert response.status_code == 200
        data = response.json()
        assert 'id' in data
        assert 'status' in data
        assert 'total_rows' in data

    def test_all_invalid_rows(self, auth_client, sample_xlsx_all_invalid):
        """全部无效行 → imported_count=0。"""
        with open(sample_xlsx_all_invalid, 'rb') as f:
            resp = auth_client.post(reverse('import-upload'), {'file': f}, format='multipart')
        session_id = resp.json()['session_id']

        response = auth_client.post(reverse('import-confirm', kwargs={'pk': session_id}))
        assert response.status_code == 200
        assert response.json()['imported_count'] == 0

    def test_imported_data_accuracy(self, auth_client, admin_user, tmp_path):
        """验证导入数据准确性。"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['标题', '描述', '优先级'])
        ws.append(['精确任务', '详细描述', '高'])
        path = tmp_path / 'accuracy.xlsx'
        wb.save(str(path))
        wb.close()

        with open(path, 'rb') as f:
            resp = auth_client.post(reverse('import-upload'), {'file': f}, format='multipart')
        session_id = resp.json()['session_id']

        auth_client.post(reverse('import-confirm', kwargs={'pk': session_id}))
        task = Task.objects.filter(title='精确任务').first()
        assert task is not None
        assert task.description == '详细描述'
        assert task.priority == 'HIGH'
        assert task.source == 'IMPORT:xlsx'
