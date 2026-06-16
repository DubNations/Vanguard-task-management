"""导出模块测试 — Excel/CSV/PDF 导出。"""
import uuid

import pytest
from django.urls import reverse

from apps.exports.models import ExportJob
from apps.tasks.models import Task
from apps.tasks.services.task_service import TaskService


@pytest.mark.django_db
class TestExportCreate:
    """导出创建测试。"""

    def test_export_excel(self, auth_client, admin_user, media_root):
        """导出 EXCEL → row_count 正确，file_name 以 .xlsx 结尾。"""
        TaskService.create_task({'title': '任务A'}, admin_user)
        TaskService.create_task({'title': '任务B'}, admin_user)

        response = auth_client.post(
            reverse('export-create'),
            {'format': 'EXCEL'},
            format='json',
        )
        assert response.status_code == 200
        data = response.json()
        assert data['row_count'] == 2
        assert data['file_name'].endswith('.xlsx')

    def test_export_csv(self, auth_client, admin_user, media_root):
        """导出 CSV → file_name 以 .csv 结尾。"""
        TaskService.create_task({'title': 'CSV任务'}, admin_user)

        response = auth_client.post(
            reverse('export-create'),
            {'format': 'CSV'},
            format='json',
        )
        assert response.status_code == 200
        data = response.json()
        assert data['file_name'].endswith('.csv')

    def test_export_pdf_unavailable(self, auth_client, admin_user, media_root):
        """导出 PDF(WeasyPrint 不可用) → 500。"""
        TaskService.create_task({'title': 'PDF任务'}, admin_user)

        response = auth_client.post(
            reverse('export-create'),
            {'format': 'PDF'},
            format='json',
        )
        assert response.status_code == 500
        assert 'WeasyPrint' in response.json()['error']

    def test_export_unsupported_format(self, auth_client, media_root):
        """不支持的格式 → 400。"""
        response = auth_client.post(
            reverse('export-create'),
            {'format': 'XML'},
            format='json',
        )
        assert response.status_code == 400

    def test_export_with_status_filter(self, auth_client, admin_user, media_root):
        """按状态筛选导出。"""
        t1 = TaskService.create_task({'title': '待领取任务'}, admin_user)
        t2 = TaskService.create_task({'title': '进行中任务'}, admin_user)
        TaskService.transition_status(t2, 'IN_PROGRESS', admin_user)

        response = auth_client.post(
            reverse('export-create'),
            {'format': 'EXCEL', 'filters': {'status': 'PENDING'}},
            format='json',
        )
        assert response.status_code == 200
        assert response.json()['row_count'] == 1

    def test_export_with_priority_filter(self, auth_client, admin_user, media_root):
        """按优先级筛选导出。"""
        TaskService.create_task({'title': '高优', 'priority': 'HIGH'}, admin_user)
        TaskService.create_task({'title': '低优', 'priority': 'LOW'}, admin_user)

        response = auth_client.post(
            reverse('export-create'),
            {'format': 'EXCEL', 'filters': {'priority': 'HIGH'}},
            format='json',
        )
        assert response.status_code == 200
        assert response.json()['row_count'] == 1

    def test_export_with_assignee_filter(self, auth_client, admin_user, regular_user, media_root):
        """按负责人筛选导出。"""
        TaskService.create_task({'title': '管理员任务'}, admin_user)
        TaskService.create_task({'title': '成员任务', 'assignee': regular_user}, admin_user)

        response = auth_client.post(
            reverse('export-create'),
            {'format': 'EXCEL', 'filters': {'assignee': str(regular_user.pk)}},
            format='json',
        )
        assert response.status_code == 200
        assert response.json()['row_count'] == 1

    def test_export_empty_dataset(self, auth_client, media_root):
        """空数据集导出 → row_count=0。"""
        response = auth_client.post(
            reverse('export-create'),
            {'format': 'EXCEL'},
            format='json',
        )
        assert response.status_code == 200
        assert response.json()['row_count'] == 0


@pytest.mark.django_db
class TestExportDownload:
    """导出下载测试。"""

    def test_download_completed(self, auth_client, admin_user, media_root):
        """下载已完成的导出 → 200，有文件内容。"""
        TaskService.create_task({'title': '下载测试任务'}, admin_user)

        resp = auth_client.post(
            reverse('export-create'),
            {'format': 'EXCEL'},
            format='json',
        )
        job_id = resp.json()['id']

        response = auth_client.get(reverse('export-download', kwargs={'pk': job_id}))
        assert response.status_code == 200
        assert b'PK' in b''.join(response.streaming_content)  # xlsx starts with PK zip header

    def test_download_nonexistent(self, auth_client, media_root):
        """下载不存在的导出 → 404。"""
        response = auth_client.get(reverse('export-download', kwargs={'pk': uuid.uuid4()}))
        assert response.status_code == 404

    def test_download_incomplete(self, auth_client, admin_user, media_root):
        """下载未完成的导出 → 404。"""
        job = ExportJob.objects.create(
            requester=admin_user,
            format='EXCEL',
            status=ExportJob.Status.PENDING,
        )
        response = auth_client.get(reverse('export-download', kwargs={'pk': job.pk}))
        assert response.status_code == 404


@pytest.mark.django_db
class TestExportHistory:
    """导出历史测试。"""

    def test_export_history_list(self, auth_client, admin_user, media_root):
        """导出历史列表 → 200。"""
        TaskService.create_task({'title': '历史任务'}, admin_user)
        auth_client.post(reverse('export-create'), {'format': 'EXCEL'}, format='json')

        response = auth_client.get(reverse('export-history'))
        assert response.status_code == 200
        data = response.json()
        assert len(data) >= 1
        assert len(data) <= 20

    def test_excel_content_headers(self, auth_client, admin_user, media_root):
        """EXCEL 内容验证 → 正确的表头。"""
        from openpyxl import load_workbook
        import io

        TaskService.create_task({'title': '表头验证任务'}, admin_user)

        resp = auth_client.post(
            reverse('export-create'),
            {'format': 'EXCEL'},
            format='json',
        )
        job_id = resp.json()['id']

        download_resp = auth_client.get(reverse('export-download', kwargs={'pk': job_id}))
        content = b''.join(download_resp.streaming_content)

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert '任务编号' in headers
        assert '标题' in headers
        assert '状态' in headers
        wb.close()

    def test_csv_encoding_bom(self, auth_client, admin_user, media_root):
        """CSV 编码 → UTF-8 BOM。"""
        TaskService.create_task({'title': 'BOM测试'}, admin_user)

        resp = auth_client.post(
            reverse('export-create'),
            {'format': 'CSV'},
            format='json',
        )
        job_id = resp.json()['id']

        download_resp = auth_client.get(reverse('export-download', kwargs={'pk': job_id}))
        content = b''.join(download_resp.streaming_content)
        assert content[:3] == b'\xef\xbb\xbf'

    def test_export_as_member(self, member_client, admin_user, regular_user, media_root):
        """以 MEMBER 身份导出 → 403（仅组长及以上可导出）。"""
        TaskService.create_task({'title': '管理员的任务'}, admin_user)

        response = member_client.post(
            reverse('export-create'),
            {'format': 'EXCEL'},
            format='json',
        )
        assert response.status_code == 403
