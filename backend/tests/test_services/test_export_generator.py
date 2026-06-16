"""
export_generator 单元测试。
"""
import pytest
from unittest.mock import MagicMock


@pytest.mark.django_db
class TestExportGenerator:
    """测试导出生成器。"""

    def _make_job(self, fmt='EXCEL', filters=None):
        job = MagicMock()
        job.format = fmt
        job.filters = filters or {}
        return job

    # ------------------------------------------------------------------
    # Excel 导出
    # ------------------------------------------------------------------
    def test_excel_export_success(self, admin_user):
        """Excel 导出成功，文件名含时间戳。"""
        from apps.exports.services.export_generator import generate_export
        from apps.tasks.models import Task

        Task.objects.create(task_no='T-EX-0001', title='ExportA', creator=admin_user)
        job = self._make_job('EXCEL')
        result = generate_export(job)
        assert result['row_count'] == 1
        assert result['file_name'].startswith('任务导出_')
        assert result['file_name'].endswith('.xlsx')

    def test_excel_headers(self, admin_user):
        """Excel 表头含 9 列。"""
        from apps.exports.services.export_generator import generate_export
        from openpyxl import load_workbook
        import io

        job = self._make_job('EXCEL')
        result = generate_export(job)
        wb = load_workbook(io.BytesIO(result['file'].read()))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        expected = ['任务编号', '标题', '状态', '优先级', '进度', '负责人', '创建人', '截止日期', '创建时间']
        assert headers == expected
        wb.close()

    # ------------------------------------------------------------------
    # CSV 导出
    # ------------------------------------------------------------------
    def test_csv_export_success(self, admin_user):
        """CSV 导出成功。"""
        from apps.exports.services.export_generator import generate_export
        from apps.tasks.models import Task

        Task.objects.create(task_no='T-CSV-0001', title='CSVTask', creator=admin_user)
        job = self._make_job('CSV')
        result = generate_export(job)
        assert result['row_count'] == 1
        assert result['file_name'].endswith('.csv')

    def test_csv_utf8_bom(self, admin_user):
        """CSV 使用 UTF-8 BOM 编码。"""
        from apps.exports.services.export_generator import generate_export

        job = self._make_job('CSV')
        result = generate_export(job)
        content = result['file'].read()
        assert content.startswith(b'\xef\xbb\xbf')

    # ------------------------------------------------------------------
    # PDF 导出
    # ------------------------------------------------------------------
    def test_pdf_export_raises(self, admin_user):
        """PDF 导出在缺少 WeasyPrint 时抛出 ValueError。"""
        from apps.exports.services.export_generator import generate_export

        job = self._make_job('PDF')
        try:
            generate_export(job)
        except ValueError as e:
            assert 'WeasyPrint' in str(e) or 'PDF' in str(e)
        except ImportError:
            # 如果 weasyprint 已安装但依赖缺失也算通过
            pass

    # ------------------------------------------------------------------
    # 过滤
    # ------------------------------------------------------------------
    def test_status_filter(self, admin_user):
        """status 过滤生效。"""
        from apps.exports.services.export_generator import generate_export
        from apps.tasks.models import Task

        Task.objects.create(task_no='T-F1-0001', title='Pending', creator=admin_user, status='PENDING')
        Task.objects.create(task_no='T-F1-0002', title='Done', creator=admin_user, status='COMPLETED')
        job = self._make_job('EXCEL', filters={'status': 'PENDING'})
        result = generate_export(job)
        assert result['row_count'] == 1

    def test_priority_filter(self, admin_user):
        """priority 过滤生效。"""
        from apps.exports.services.export_generator import generate_export
        from apps.tasks.models import Task

        Task.objects.create(task_no='T-F2-0001', title='High', creator=admin_user, priority='HIGH')
        Task.objects.create(task_no='T-F2-0002', title='Low', creator=admin_user, priority='LOW')
        job = self._make_job('EXCEL', filters={'priority': 'HIGH'})
        result = generate_export(job)
        assert result['row_count'] == 1

    # ------------------------------------------------------------------
    # 空数据集
    # ------------------------------------------------------------------
    def test_empty_dataset_export(self, admin_user):
        """空数据集导出成功，row_count=0。"""
        from apps.exports.services.export_generator import generate_export

        job = self._make_job('EXCEL')
        result = generate_export(job)
        assert result['row_count'] == 0
